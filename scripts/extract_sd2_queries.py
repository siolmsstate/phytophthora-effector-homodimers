from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "41467_2024_54452_MOESM4_ESM.xlsx"
OUTDIR = ROOT / "outputs" / "sd2_colabfold"
MANIFEST = OUTDIR / "sd2_homodimer_manifest.csv"
QUERY_CSV = OUTDIR / "sd2_homodimer_queries.csv"

VALID_AA = set("ACDEFGHIKLMNPQRSTVWY")


def sanitize_identifier(value: object, fallback: str) -> str:
    text = str(value or "").strip() or fallback
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_.-")
    return text or fallback


def clean_sequence(seq: object) -> tuple[str, str]:
    original = str(seq or "").strip()
    cleaned = re.sub(r"\s+", "", original).upper().replace("-", "")
    invalid = "".join(sorted(set(cleaned) - VALID_AA))
    if invalid:
        raise ValueError(f"Unsupported amino acid code(s) after cleaning: {invalid}")
    return original, cleaned


def main() -> None:
    OUTDIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["SD2"]

    headers = [cell.value for cell in ws[2]]
    col = {name: idx for idx, name in enumerate(headers)}
    required = ["List_No", "BLASTP matched", "Working name", "Phylogroup", "Cloned_AA"]
    missing = [name for name in required if name not in col]
    if missing:
        raise RuntimeError(f"Missing expected SD2 column(s): {', '.join(missing)}")

    records = []
    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not any(row):
            continue
        list_no = row[col["List_No"]]
        if list_no in (None, ""):
            continue
        original, cleaned = clean_sequence(row[col["Cloned_AA"]])
        if not cleaned:
            continue
        name = sanitize_identifier(row[col["Working name"]], f"row_{excel_row_idx}")
        job_id = f"SD2_{int(list_no):03d}_{name}"
        removed_chars = "".join(sorted(set(str(original)) - set(cleaned)))
        records.append(
            {
                "id": job_id,
                "list_no": int(list_no),
                "source_excel_row": excel_row_idx,
                "blastp_matched": row[col["BLASTP matched"]] or "",
                "working_name": row[col["Working name"]] or "",
                "phylogroup": row[col["Phylogroup"]] or "",
                "original_sequence": original,
                "cleaned_sequence": cleaned,
                "homodimer_sequence": f"{cleaned}:{cleaned}",
                "monomer_length": len(cleaned),
                "complex_length": len(cleaned) * 2,
                "removed_chars": removed_chars,
            }
        )

    if len(records) != 74:
        raise RuntimeError(f"Expected 74 SD2 sequences, found {len(records)}")

    with MANIFEST.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0]))
        writer.writeheader()
        writer.writerows(records)

    with QUERY_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["id", "sequence"])
        writer.writeheader()
        writer.writerows({"id": r["id"], "sequence": r["homodimer_sequence"]} for r in records)

    print(f"Wrote {len(records)} homodimer queries")
    print(MANIFEST)
    print(QUERY_CSV)


if __name__ == "__main__":
    main()
